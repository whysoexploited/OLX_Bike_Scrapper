import os
import time
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

# --- Setup ---
search_term = input("Enter the bike model to search (e.g. Yamaha MT-07): ").strip()
search_slug = search_term.lower().replace(" ", "-")

SEEN_ADS_FILE = f"seen_{search_slug}.txt"
EXCEL_FILE = f"{search_slug}_ads.xlsx"

# Headless Chrome setup
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# --- Helper: Ad still active ---
def is_ad_still_active(url):
    try:
        response = requests.get(url, timeout=10)
        if response.status_code != 200 or "Anuntul pe care il cauti nu mai exista" in response.text:
            return False
        return True
    except:
        return False

# --- Load existing ads ---
if os.path.exists(SEEN_ADS_FILE):
    with open(SEEN_ADS_FILE, "r") as f:
        seen_urls = set(f.read().splitlines())
else:
    seen_urls = set()

# --- Load previous prices if available ---
if os.path.exists(EXCEL_FILE):
    df_existing = pd.read_excel(EXCEL_FILE)
    old_prices = dict(zip(df_existing["URL"], df_existing["Price"]))
else:
    df_existing = pd.DataFrame(columns=["Title", "URL", "Price", "Location"])
    old_prices = {}

# --- Scrap Multiple Pages ---
all_ads = []
for page in range(1, 6):  # Scrape first 5 pages
    search_url = f"https://www.olx.ro/d/oferte/q-{search_slug}/?page={page}"
    print(f"Scraping page {page}: {search_url}")
    driver.get(search_url)
    time.sleep(3)

    ads = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="listing-grid"] > div')
    if not ads:
        break

    for ad in ads:
        try:
            link_element = ad.find_element(By.CSS_SELECTOR, 'a.css-1tqlkj0')
            url = link_element.get_attribute("href").strip()

            title = link_element.get_attribute("title") or link_element.text.strip() or "No title"
            if url in seen_urls:
                continue

            price_elem = ad.find_element(By.CSS_SELECTOR, '[data-testid="ad-price"]')
            price = price_elem.text.strip()

            location_elem = ad.find_element(By.CSS_SELECTOR, '[data-testid="location-date"]')
            location = location_elem.text.split(" - ")[0].strip()

            old_price = old_prices.get(url)
            all_ads.append({
                "Title": title,
                "URL": url,
                "Price": price,
                "Location": location,
                "Previous Price": old_price if old_price and old_price != price else None
            })
        except Exception:
            continue

driver.quit()

# --- Clean removed ads from Excel ---
if not df_existing.empty:
    df_existing = df_existing[df_existing["URL"].apply(is_ad_still_active)]
    seen_urls = set(df_existing["URL"])

# --- Add new ads ---
new_ads = [ad for ad in all_ads if ad["URL"] not in seen_urls]
if new_ads:
    print(f"Found {len(new_ads)} new ads:")
    for ad in new_ads:
        print(f"{ad['Title']} - {ad['Price']} - {ad['Location']}\n{ad['URL']}\n")

    df_new = pd.DataFrame(new_ads)
    df_result = pd.concat([df_existing, df_new], ignore_index=True)

    # Save to Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df_result.to_excel(writer, index=False, sheet_name='Ads')
        workbook = writer.book
        sheet = writer.sheets['Ads']

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for idx, row in df_result.iterrows():
            if pd.notna(row.get("Previous Price")):
                for col in range(1, len(df_result.columns) + 1):
                    sheet.cell(row=idx + 2, column=col).fill = green_fill

    # Save new seen URLs
    with open(SEEN_ADS_FILE, "w") as f:
        for url in df_result["URL"]:
            f.write(url + "\n")
else:
    print("No new ads found.")
